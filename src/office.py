from __future__ import annotations

import re
import shutil
import win32com.client as win32
from pathlib import Path
from datetime import datetime
from typing import Any, Mapping
from openpyxl import load_workbook
from docx import Document
import docx2pdf
from pypdf import PdfReader, PdfWriter
from scraper import Scraper
from data import DataManager
from settings import SettingsManager


class OfficeDocumentManager:
    def __init__(self, base_dir: str | Path | None = None) -> None:
        self.base_dir = Path(base_dir) if base_dir else Path(__file__).resolve().parent
        self.templates_dir = self.base_dir / "templates"
        self.documents_dir = self.base_dir / "documents"
        self.alias_mapping = {}

    def complete_pricing_spreadsheet(self, data: DataManager, variables) -> Path:

        variables.update(data)
        workbook = load_workbook(data.files.pricing_spreadsheet_template)

        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        for placeholder, replacement in variables.get_all().items():
                            if placeholder in str(cell.value):
                                try:
                                    int(replacement)
                                    cell.value = int(
                                        cell.value.replace(
                                            placeholder, str(replacement)
                                        )
                                    )
                                except (ValueError, TypeError):
                                    cell.value = cell.value.replace(
                                        placeholder, str(replacement)
                                    )
                                    
        workbook.save(data.files.pricing_spreadsheet)
        return data.files.pricing_spreadsheet

    def complete_cover_letter(self, data: DataManager, variables) -> Path:
        data.update_from_pricing_spreadsheet()
        variables.update(data)
        document = Document(str(data.files.cover_letter_template))

        for paragraph in document.paragraphs:
            self._replace_in_paragraph(paragraph, variables.get_all())

        document.save(str(data.files.cover_letter_docx))
        return data.files.cover_letter_docx

    def finalize_proposal(self, data: DataManager, variables, scraper: Scraper) -> Path:
        docx2pdf.convert(
            str(data.files.cover_letter_docx), str(data.files.cover_letter_pdf)
        )

        writer = PdfWriter()
        writer.append(str(data.files.cover_letter_pdf))
        writer.append(str(data.files.solargraf_proposal))
        
        specsheet_pages = {}

        if data.files.settings.is_file():
            workbook = load_workbook(data.files.settings, data_only=True, read_only=True)
            sheet = workbook["Spec Sheets"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                model, value = row[0], row[1]
                if model and value is not None:
                    if model == variables.resolve(r'{{PANEL_MODEL}}'):
                        specsheet_pages['panel'] = int(value)
                    elif model == variables.resolve(r'{{INVERTER_MODEL}}'):
                        specsheet_pages['inverter'] = int(value)
                    elif model == variables.resolve(r'{{BATTERY_MODEL}}'):
                        specsheet_pages['battery'] = int(value)
            workbook.close()
        
        data.files.specsheets_folder.mkdir(parents=True, exist_ok=True)

        def download_and_trim(url: str, specsheet_name: str) -> None:
            specsheet_path = scraper.download(
                url, output_dir=data.files.specsheets_folder
            )
            max_pages = specsheet_pages.get(specsheet_name)
            if max_pages is None:
                return

            reader = PdfReader(str(specsheet_path))
            writer = PdfWriter()
            for page in reader.pages[:max_pages]:
                writer.add_page(page)

            with open(specsheet_path, "wb") as output_file:
                writer.write(output_file)

        if data.system.type == 0:
            download_and_trim(data.system.panel.specsheet_url, "panel")
            download_and_trim(data.system.inverter.specsheet_url, "inverter")
        if data.system.type == 1:
            download_and_trim(data.system.battery.specsheet_url, "battery")
        if data.system.type == 2:
            download_and_trim(data.system.panel.specsheet_url, "panel")
            download_and_trim(data.system.inverter.specsheet_url, "inverter")
            download_and_trim(data.system.battery.specsheet_url, "battery")

        for specsheet_file in data.files.specsheets_folder.glob("*.pdf"):
            writer.append(str(specsheet_file))
        writer.write(str(data.files.final_proposal))
        return data.files.final_proposal

    def generate_email(self, data: DataManager, variables) -> None:
        data.update_from_pricing_spreadsheet()
        variables.update(data)

        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)  # 0: olMailItem
        mail.Subject = (
            f"Mid-State Solar - Proposal for {variables.resolve(r'{{SYSTEM_TYPE}}')}"
        )
        mail.To = variables.resolve(r"{{CLIENT_EMAIL}}")

        email_body = data.files.email_template.read_text(encoding="utf-8")
        for placeholder, replacement in variables.get_all().items():
            email_body = email_body.replace(str(placeholder), str(replacement))

        mail.HtmlBody = email_body
        mail.Attachments.Add(str(data.files.final_proposal))
        mail.Display(False)

    @staticmethod
    def _replace_in_paragraph(paragraph, replacements: Mapping[str, Any]) -> None:
        runs = paragraph.runs
        if not runs:
            return

        full_text = "".join(run.text for run in runs)
        if not full_text:
            return

        # Longest keys first, so a key that's a prefix of another (e.g. {{NAME}}
        # vs {{NAME_FULL}}) can't accidentally swallow part of the longer one.
        keys = sorted((str(k) for k in replacements if str(k)), key=len, reverse=True)
        if not keys:
            return
        pattern = re.compile("|".join(re.escape(k) for k in keys))

        matches = list(pattern.finditer(full_text))
        if not matches:
            return  # nothing to do, don't touch run formatting

        # Offset of each run within the concatenated full_text.
        run_spans = []
        pos = 0
        for run in runs:
            run_spans.append((pos, pos + len(run.text)))
            pos += len(run.text)

        # Flatten full_text into segments: unchanged text and matched placeholders.
        segments = []  # (seg_start, seg_end, is_match, replacement_or_None)
        cursor = 0
        for m in matches:
            if m.start() > cursor:
                segments.append((cursor, m.start(), False, None))
            segments.append((m.start(), m.end(), True, str(replacements[m.group(0)])))
            cursor = m.end()
        if cursor < len(full_text):
            segments.append((cursor, len(full_text), False, None))

        # Rebuild each run from whichever segments overlap its original span.
        for run, (rs, re_) in zip(runs, run_spans):
            if rs == re_:
                continue  # empty run

            pieces = []
            for seg_start, seg_end, is_match, replacement in segments:
                if seg_end <= rs or seg_start >= re_:
                    continue  # segment doesn't touch this run

                if not is_match:
                    pieces.append(full_text[max(seg_start, rs):min(seg_end, re_)])
                else:
                    # Emit the replacement only in the run that contains the
                    # *start* of the match, so a placeholder spanning several
                    # runs doesn't get duplicated.
                    if rs <= seg_start < re_:
                        pieces.append(replacement)

            run.text = "".join(pieces)

    @staticmethod
    def get_cell_value(data: DataManager, sheet_name: str, cell_ref: str) -> Any:
        if data.files.pricing_spreadsheet.exists():
            try:
                wb = load_workbook(
                    data.files.pricing_spreadsheet, data_only=True, read_only=True
                )
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    cell = sheet[cell_ref]
                    return cell.value
            except Exception as e:
                print(f"Error reading cell {cell_ref} from sheet {sheet_name}: {e}")
                return 0
        return 0

    def copy_to_fileserver(self, data: DataManager, settings: SettingsManager) -> Path:
        if not settings.files.bids_folder:
            raise ValueError("FILESERVER_BIDS_FOLDER environment variable is not set.")

        destination_folder = (
            Path(settings.files.bids_folder)
            / f"{data.client.last_name}, {data.client.first_name}"
        )
        destination_folder.mkdir(parents=True, exist_ok=True)

        shutil.copy(
            data.files.final_proposal,
            destination_folder
            / f"{datetime.now().strftime('%y_%m%d')} {data.client.last_name}, {data.client.first_name} {"PV" if data.system.type == 0 else "ESS" if data.system.type == 1 else "PV ESS"} Proposal.pdf",
        )
        shutil.copy(
            data.files.pricing_spreadsheet,
            destination_folder
            / f"{datetime.now().strftime('%y_%m%d')} {data.client.last_name}, {data.client.first_name} {"PV" if data.system.type == 0 else "ESS" if data.system.type == 1 else "PV ESS"} Bid.xlsx",
        )
        shutil.copy(
            data.files.cover_letter_docx,
            destination_folder
            / f"PV ESS Proposal - {data.client.last_name}, {data.client.first_name}.docx",
        )
        shutil.copy(
            data.files.solargraf_proposal,
            destination_folder / f"Proposal for - {data.client.name}.pdf",
        )
        shutil.copytree(
            data.files.images_folder, destination_folder / "images", dirs_exist_ok=True
        )

        return destination_folder
