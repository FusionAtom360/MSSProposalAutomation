import datetime

from data import DataManager
from openpyxl import load_workbook
from datetime import datetime

class TemplateManager:
    def __init__(self, data: DataManager):
        self.templates = {}
        self.aliases = {}
        self.update(data)

    def update(self, data: DataManager):
        self._update_from_settings_spreadsheet(data)
        self.templates[r"{{JOB_NAME}}"] = data.project.name.upper()
        self.templates[r"{{CLIENT_NAME}}"] = data.client.name
        self.templates[r"{{CLIENT_FIRST_NAME}}"] = data.client.first_name
        self.templates[r"{{CLIENT_LAST_NAME}}"] = data.client.last_name
        self.templates[r"{{CLIENT_PHONE}}"] = data.client.phone
        self.templates[r"{{CLIENT_EMAIL}}"] = data.client.email
        self.templates[r"{{CLIENT_ADDRESS_1}}"] = data.client.address.street
        self.templates[r"{{CLIENT_ADDRESS_2}}"] = (
            f"{data.client.address.city}, {data.client.address.region_code} {data.client.address.postal_code}"
        )
        self.templates[r"{{CLIENT_ADDRESS_STREET}}"] = data.client.address.street
        self.templates[r"{{CLIENT_ADDRESS_CITY}}"] = data.client.address.city
        self.templates[r"{{CLIENT_ADDRESS_ZIP}}"] = data.client.address.postal_code

        self.templates[r"{{UTILITY_NAME}}"] = self.aliases.get(data.utility.name, data.utility.name)
        self.templates[r"{{ANNUAL_CONSUMPTION}}"] = data.utility.annual_consumption

        self.templates[r"{{PANEL_MANUFACTURER}}"] = self.aliases.get(data.system.panel.manufacturer, data.system.panel.manufacturer)
        self.templates[r"{{PANEL_MANUFACTURER_UPPERCASE}}"] = self.aliases.get(data.system.panel.manufacturer, data.system.panel.manufacturer).upper()
        self.templates[r"{{PANEL_MODEL}}"] = self.aliases.get(data.system.panel.model, data.system.panel.model)
        self.templates[r"{{PANEL_TYPE}}"] = (
            "monocrystalline" if data.system.panel.type == "mono"
            else "polycrystalline" if data.system.panel.type == "poly"
            else "bifacial"
        )
        self.templates[r"{{PANEL_SIZE}}"] = data.system.panel.size
        self.templates[r"{{PANEL_PTC}}"] = data.system.panel.ptc
        self.templates[r"{{PANEL_PTC_RATIO}}"] = round(
            data.system.panel.ptc / data.system.panel.size * 100
        ) if data.system.panel.size and data.system.panel.ptc else None
        self.templates[r"{{PANEL_WARRANTY}}"] = data.system.panel.warranty
        self.templates[r"{{PANEL_COUNT}}"] = data.system.panel.count

        self.templates[r"{{INVERTER_MANUFACTURER}}"] = self.aliases.get(data.system.inverter.manufacturer, data.system.inverter.manufacturer)
        self.templates[r"{{INVERTER_MANUFACTURER_UPPERCASE}}"] = self.aliases.get(data.system.inverter.manufacturer, data.system.inverter.manufacturer).upper()
        self.templates[r"{{INVERTER_MODEL}}"] = self.aliases.get(data.system.inverter.model, data.system.inverter.model)
        self.templates[r"{{INVERTER_WARRANTY}}"] = data.system.inverter.warranty
        self.templates[r"{{INVERTER_COUNT}}"] = data.system.inverter.count
        self.templates[r"{{INVERTER_EFFICIENCY}}"] = f"{round(data.system.inverter.efficiency*100)}" if data.system.inverter.efficiency else None

        self.templates[r"{{BATTERY_MANUFACTURER}}"] = self.aliases.get(data.system.battery.manufacturer, data.system.battery.manufacturer)
        self.templates[r"{{BATTERY_MANUFACTURER_UPPERCASE}}"] = self.aliases.get(data.system.battery.manufacturer, data.system.battery.manufacturer).upper()
        self.templates[r"{{BATTERY_MODEL}}"] = self.aliases.get(data.system.battery.model, data.system.battery.model)
        self.templates[r"{{BATTERY_CAPACITY}}"] = data.system.battery.capacity
        self.templates[r"{{BATTERY_WARRANTY}}"] = data.system.battery.warranty
        self.templates[r"{{BATTERY_WARRANTY_CYCLES}}"] = data.system.battery.warranty * 400 if data.system.battery.warranty else None
        self.templates[r"{{BATTERY_WARRANTY_CAPACITY}}"] = 60
        self.templates[r"{{BATTERY_COUNT}}"] = data.system.battery.count

        self.templates[r"{{SYSTEM_TYPE}}"] = "PV" if data.system.type == 0 else "ESS" if data.system.type == 1 else "PV & ESS"
        self.templates[r"{{PV_SIZE}}"] = data.system.pv_size
        self.templates[r"{{ESS_SIZE}}"] = data.system.ess_size
        self.templates[r"{{ESS_TYPE}}"] = self.resolve(r"{{ESS_TYPE_ON_GRID}}") if data.system.ess_type == 0 else self.resolve(r"{{ESS_TYPE_OFF_GRID}}")
        self.templates[r"{{OFF_GRID_SUBPANEL}}"] = self.resolve(r"{{OFF_GRID_SUBPANEL_OPTION}}") if data.system.ess_type == 0 else self.resolve(r"{{OFF_GRID_SUBPANEL_OMIT}}")

        self.templates[r"{{PREPARED_DATE_MMDDYYYY}}"] = datetime.now().strftime("%m/%d/%Y")
        self.templates[r"{{ENPHASE_PLATINUM_INSTALLER}}"] = (
            self.resolve(r"{{ENPHASE_PLATINUM_INSTALLER}}")
            if data.system.inverter.manufacturer == "Enphase Energy Inc."
            else ""
        )
        self.templates[r"{{TREE_TRIMMING_REQUIRED}}"] = self.resolve(r"{{TREE_TRIMMING_REQUIRED}}") if data.system.tree_trimming_required else ""
        self.templates[r"{{UTILITY_NEM}}"] = self.resolve(r"{{UTILITY_NEM_MID}}") if data.utility.name == "Modesto Irrigation District" else self.resolve(r"{{UTILITY_NEM_PGE}}")
        self.templates[r"{{UTILITY_APPLICATION_DESCRIPTION}}"] = self.resolve(r"{{UTILITY_APPLICATION_DESCRIPTION_MID}}") if data.utility.name == "Modesto Irrigation District" else self.resolve(r"{{UTILITY_APPLICATION_DESCRIPTION_PGE}}")
        self.templates[r"{{BATTERY_PLURAL}}"] = "Batteries" if data.system.battery.count > 1 else "Battery"
        self.templates[r"{{UTILITY_LINK}}"] = self.resolve(r"{{UTILITY_LINK_MID}}") if data.utility.name == "Modesto Irrigation District" else self.resolve(r"{{UTILITY_LINK_PGE}}")
        self.templates[r"{{APPLICATION_FEE}}"] = self.resolve(r'{{APPLICATION_FEE_MID}}') if data.utility.name == "Modesto Irrigation District" else self.resolve(r'{{APPLICATION_FEE_PGE}}')

        self.templates[r"{{PV_COST}}"] = f"${int(data.pricing.pv_cost):,}"
        self.templates[r"{{ESS_COST}}"] = f"${int(data.pricing.ess_cost):,}"
        self.templates[r"{{TOTAL_COST}}"] = f"${int(data.pricing.total_cost):,}"
        self.templates[r"{{LOAN_TERM}}"] = data.pricing.loan_term
        self.templates[r"{{LOAN_INTEREST_RATE}}"] = f"{data.pricing.loan_interest_rate*100:.2f}"
        self.templates[r"{{LOAN_MONTHLY_PAYMENT}}"] = f"${int(data.pricing.loan_monthly_payment):,}"
    
    def _update_from_settings_spreadsheet(self, data: DataManager):
        if data.files.settings.is_file():
            workbook = load_workbook(data.files.settings, data_only=True, read_only=True)
            sheet = workbook["Aliases"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                placeholder, value = row[0], row[1]
                if placeholder and value is not None:
                    self.aliases[str(placeholder)] = str(value)
    
            sheet = workbook["Variables"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                placeholder, value = row[0], row[1]
                if placeholder and value is not None:
                    self.templates[str(placeholder)] = str(value)
    

    def get_all(self):
        return self.templates

    def resolve(self, template_name):
        if self.templates.get(template_name, None) is None:
            return template_name
        return self.templates.get(template_name, template_name)
