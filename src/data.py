import json
from dataclasses import dataclass
import math
import os
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from settings import SettingsManager
from datetime import datetime


class DataManager:
    def __init__(self):
        self.data = {}
        self.settings = SettingsManager()
        self.project = self.Project()
        self.client = self.Client()
        self.utility = self.Utility()
        self.system = self.System()
        self.pricing = self.Pricing()
        self.files = self.Files()
        self.references = self.References()

    def load_json(self, json_data: dict, save_to_file: bool = True):
        self.data = json_data
        self.project.id = self._get("id", default=0)
        self.project.api_id = self._get("public_id", default="")
        self.files.modules_spreadsheet = (
                    self.settings.files.templates_folder / "modules.xlsx"
                )
        
        self.client.first_name = (
            self._get("client_name", default="").split()[0]
            if self._get("client_name", default="")
            else ""
        )
        self.client.last_name = (
            " ".join(self._get("client_name", default=". .").split()[1:])
            if self._get("client_name", default="")
            else ""
        )
        self.client.name = self._get("client_name", default="")
        self.client.phone = self._get("client_phone", default="")
        self.client.email = self._get("client_email", default="")
        self.client.address.street = self._get("Position", "street", default="")
        self.client.address.city = self._get("Position", "city", default="")
        self.client.address.region_code = self._get(
            "Position", "region_code", default=""
        )
        self.client.address.postal_code = self._get(
            "Position", "postal_code", default=""
        )

        self.project.name = self.client.name.upper()

        self.system.type = self._get_system_type()
        self.system.pv_size = self._get("proposals", 0, "sizeInKw")
        self.system.panel.manufacturer = self._get(
            "Materials", "panel", 0, "Manufacturer", "name"
        )
        self.system.panel.model = self._get("Materials", "panel", 0, "name")
        self.system.panel.type = self._get("Materials", "panel", 0, "type")
        self.system.panel.size = self._get("Materials", "panel", 0, "size_in_watts")
        self.system.panel.ptc = self._get_panel_ptc()
        self.system.panel.warranty = self._get("Materials", "panel", 0, "warranty")
        self.system.panel.count = self._get("Materials", "panel", 0, "count", default=0)
        self.system.panel.specsheet_url = self._get(
            "Materials", "panel", 0, "spec_sheet"
        )

        self.utility.name = self._get(
            "Settings",
            "Pricing",
            "ProjectConnection",
            "genability_lse_name",
            default="",
        )
        self.utility.annual_consumption = self._calculate_annual_consumption(
            self._get("Settings", "Pricing", "ProjectConnection", "Bills", default="")
        )

        self.system.inverter.manufacturer = self._get(
            "Materials", "inverter", 0, "Manufacturer", "name"
        )
        self.system.inverter.model = self._get("Materials", "inverter", 0, "name")
        self.system.inverter.warranty = self._get(
            "Materials", "inverter", 0, "warranty"
        )
        self.system.inverter.count = self._get(
            "Materials", "inverter", 0, "count", default=0
        )
        self.system.inverter.specsheet_url = self._get(
            "Materials", "inverter", 0, "spec_sheet"
        )

        self.system.battery.manufacturer = self._get(
            "Materials", "batteryBackup", 0, "Manufacturer", "name"
        )
        self.system.battery.model = self._get("Materials", "batteryBackup", 0, "name")
        self.system.battery.capacity = self._get(
            "Materials", "batteryBackup", 0, "capacity"
        )
        self.system.battery.warranty = self._get(
            "Materials", "batteryBackup", 0, "warranty"
        )
        self.system.battery.count = self._get(
            "Materials", "batteryBackup", 0, "count", default=0
        )
        self.system.battery.specsheet_url = self._get(
            "Materials", "batteryBackup", 0, "spec_sheet"
        )

        self.system.ess_size = (
            self.system.battery.capacity * self.system.battery.count
            if self.system.battery.capacity and self.system.battery.count
            else 0.0
        )

        self.project.financial_id = self._get(
            "Settings", "Pricing", "FinancialOptions", 0, "id"
        )
        self.project.proposal_id = self._get(
            "Settings", "Pricing", "ProjectConnection", "proposal_id"
        )

        self.files.settings = self.settings.files.templates_folder / "settings.xlsx"
        self.files.pricing_spreadsheet_template = (
            self.settings.files.templates_folder / "pricing_spreadsheet.xlsx"
        )
        self.files.cover_letter_template = (
            self.settings.files.templates_folder
            / f"cover_letter_{self.system.type}.docx"
        )
        self.files.solargraf_proposal_template = (
            self.settings.files.templates_folder / "generated_proposal.docx"
        )
        self.files.email_template = self.settings.files.templates_folder / "email.html"
        self.files.project_folder = self.settings.files.projects_folder / str(
            self.project.id
        )
        self.files.images_folder = self.files.project_folder / "images"
        self.files.documents_folder = self.files.project_folder / "documents"
        self.files.specsheets_folder = self.files.documents_folder / "specsheets"
        self.files.solargraf_proposal = (
            self.files.documents_folder / "solargraf_proposal.pdf"
        )
        self.files.pricing_spreadsheet = (
            self.files.documents_folder / "pricing_spreadsheet.xlsx"
        )
        self.files.cover_letter_docx = self.files.documents_folder / "cover_letter.docx"
        self.files.cover_letter_pdf = self.files.documents_folder / "cover_letter.pdf"
        self.files.final_proposal = (
            self.files.documents_folder
            / f"{datetime.now().strftime('%y_%m%d')} {self.client.last_name}, {self.client.first_name} {"PV" if self.system.type == 0 else "ESS" if self.system.type == 1 else "PV ESS"} Proposal.pdf"
        )

        if save_to_file:
            self.files.project_folder.mkdir(parents=True, exist_ok=True)
            self.files.images_folder.mkdir(parents=True, exist_ok=True)
            self.files.documents_folder.mkdir(parents=True, exist_ok=True)
            self.files.specsheets_folder.mkdir(parents=True, exist_ok=True)

            with open(self.files.project_folder / "data.json", "w") as f:
                json.dump(self.data, f, indent=4)

    def set_api_id(self, api_id):
        self.project.api_id = api_id

    def _set_annual_production(self, annual_production):
        self.system.annual_production = annual_production

    def _get(self, *keys, default=None) -> Any:
        data = self.data
        try:
            for key in keys:
                if isinstance(key, int):
                    data = data[key]
                else:
                    data = data.get(key)

                if data is None:
                    return default
            return data
        except (KeyError, IndexError, TypeError):
            return default

    @staticmethod
    def _clean_float(value):
        try:
            return float(value)
        except (ValueError, TypeError):
            print(
                f"Warning: Unable to convert value '{value}' to float. Returning 0.0 instead."
            )
            return 0.0

    @staticmethod
    def _clean_int(value):
        try:
            return int(value)
        except (ValueError, TypeError):
            return 0

    def update_from_pricing_spreadsheet(self):
        if self.files.settings.is_file():
            workbook = load_workbook(
                self.files.settings, data_only=True, read_only=True
            )
            sheet = workbook["Spreadsheet"]
            self.references.tree_trimming_required = sheet["B2"].value
            self.references.ess_type = sheet["B3"].value
            self.references.pv_cost = sheet["B4"].value
            self.references.ess_cost = sheet["B5"].value
            self.references.total_cost = sheet["B6"].value
            self.references.loan_term = sheet["B7"].value
            self.references.loan_interest_rate = sheet["B8"].value
            self.references.loan_monthly_payment = sheet["B9"].value
            self.references.panel_ptc = sheet["B10"].value
            self.references.inverter_efficiency = sheet["B11"].value

        if self.files.pricing_spreadsheet.is_file():
            workbook = load_workbook(
                self.files.pricing_spreadsheet, data_only=True, read_only=True
            )
            sheet = workbook["Pricing Calculator"]
            self.system.tree_trimming_required = sheet[
                self.references.tree_trimming_required
            ].value
            self.system.ess_type = 0 if (sheet[self.references.ess_type].value == "Grid-Tied") else 1
            self.pricing.pv_cost = float(sheet[self.references.pv_cost].value) if sheet[self.references.pv_cost].value is not None else 0.0
            self.pricing.ess_cost = float(sheet[self.references.ess_cost].value) if sheet[self.references.ess_cost].value is not None else 0.0
            self.pricing.total_cost = float(sheet[self.references.total_cost].value) if sheet[self.references.total_cost].value is not None else 0.0
            self.pricing.loan_term = int(sheet[self.references.loan_term].value)
            self.pricing.loan_interest_rate = float(
                sheet[self.references.loan_interest_rate].value
            )
            self.pricing.loan_monthly_payment = math.ceil(
                sheet[self.references.loan_monthly_payment].value
            )
            self.system.panel.ptc = int(float(sheet[self.references.panel_ptc].value))
            self.system.inverter.efficiency = float(sheet[self.references.inverter_efficiency].value) if sheet[self.references.inverter_efficiency].value is not None else 0.97

    def _get_system_type(self):
        if self._get("is_roofing_only"):
            return 0
        elif self._get("is_battery_only"):
            return 1
        else:
            return 2

    def _calculate_annual_consumption(self, bills):
        total_consumption = 0.0
        # TODO: Rewrite for real data
        for bill in bills:
            consumption = bill.get("consumptionInKwh", 0.0)
            total_consumption += consumption
        return total_consumption

    def _get_panel_ptc(self) -> int | None:
        if self.system.panel.ptc is not None and self.system.panel.ptc > 0:
            return self.system.panel.ptc
        elif self.files.modules_spreadsheet.is_file():
            workbook = load_workbook(
                self.files.modules_spreadsheet, data_only=True, read_only=True
            )
            sheet = workbook["PV Module-Full"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                model, ptc = row[1], row[5]
                if self.system.panel.model is not None and model is not None:
                    if self.system.panel.model.lower() in str(model).lower():
                        return ptc
        return None

    @dataclass
    class Project:
        id: int = 0
        api_id: str = ""
        financial_id: str = ""
        proposal_id: str = ""
        name: str = ""

    class Client:
        def __init__(self):
            self.first_name: str = ""
            self.last_name: str = ""
            self.name: str = ""
            self.phone: str = ""
            self.email: str = ""
            self.address = self.Address()

        @dataclass
        class Address:
            street: str = ""
            city: str = ""
            region_code: str = ""
            postal_code: str = ""

    @dataclass
    class Utility:
        name: str = ""
        annual_consumption: float = 0.0

    @dataclass
    class System:
        def __init__(self):
            self.type: int = 0
            self.ess_type: int = 0
            self.pv_size: float = 0.0
            self.ess_size: float = 0.0
            self.panel = self.Panel()
            self.inverter = self.Inverter()
            self.battery = self.Battery()
            self.annual_production: float = 0.0
            self.tree_trimming_required: bool = False

        @dataclass
        class Panel:
            manufacturer: str = ""
            model: str = ""
            size: int = 0
            type: str = ""
            warranty: int = 0
            ptc: int = 0
            count: int = 0
            specsheet_url: str = ""

        @dataclass
        class Inverter:
            manufacturer: str = ""
            model: str = ""
            efficiency: float = 0.0
            warranty: int = 0
            count: int = 0
            specsheet_url: str = ""

        @dataclass
        class Battery:
            manufacturer: str = ""
            model: str = ""
            capacity: float = 0.0
            warranty: int = 0
            count: int = 0
            specsheet_url: str = ""

    @dataclass
    class Pricing:
        id: str = ""
        pv_cost = 0.0
        ess_cost: float = 0.0
        total_cost: float = 0.0
        cost_per_watt: float = 0.0
        loan_term: int = int(os.getenv("LOAN_TERM", "0"))
        loan_interest_rate: float = float(os.getenv("LOAN_INTEREST_RATE", "0.0"))
        loan_monthly_payment: float = 0

    @dataclass
    class Files:
        pricing_spreadsheet_template: Path = Path()
        cover_letter_template: Path = Path()
        solargraf_proposal_template: Path = Path()
        modules_spreadsheet: Path = Path()
        settings: Path = Path()
        email_template: Path = Path()
        project_folder: Path = Path()
        images_folder: Path = Path()
        documents_folder: Path = Path()
        specsheets_folder: Path = Path()
        solargraf_proposal: Path = Path()
        pricing_spreadsheet: Path = Path()
        cover_letter_docx: Path = Path()
        cover_letter_pdf: Path = Path()
        final_proposal: Path = Path()

    @dataclass
    class References:
        tree_trimming_required: str = ""
        ess_type: str = ""
        pv_cost: str = ""
        ess_cost: str = ""
        total_cost: str = ""
        loan_term: str = ""
        loan_interest_rate: str = ""
        loan_monthly_payment: str = ""
        panel_ptc: str = ""
        inverter_efficiency: str = ""
